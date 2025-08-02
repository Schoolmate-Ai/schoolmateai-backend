from typing import List, Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_, and_
import uuid
from datetime import datetime, timedelta
from shared.db import get_db
from shared.auth import get_current_user
from services.user_management.models.users import SchoolUser, SchoolUserRole
from services.user_management.models.classes import SchoolClass
from services.attendance_management_system.models.attendance import Attendance, AttendanceStatus
from services.attendance_management_system.schemas.attendance import (
    DailyAttendanceCreate,
    DailyAttendanceResponse,
    AttendanceOut,
    StudentAttendanceRecord,
    AttendanceStatus as SchemaAttendanceStatus
)
from fastapi.responses import FileResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, Reference
import tempfile
import os

router = APIRouter(prefix="/attendance", tags=["Attendance Management"])


async def _get_class_students(db: AsyncSession, class_id: uuid.UUID) -> List[SchoolUser]:
    result = await db.execute(
        select(SchoolUser).where(
            and_(
                SchoolUser.class_id == class_id,
                SchoolUser.role == SchoolUserRole.STUDENT,
                SchoolUser.is_active == True
            )
        )
    )
    return result.scalars().all()


# --- TAKE DAILY ATTENDANCE BY CLASS-TEACHER ---
# @router.post("/daily", response_model=DailyAttendanceResponse)
@router.post("/daily")
async def record_daily_attendance(
    attendance_data: DailyAttendanceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Record daily attendance for a class. 
    Only students with non-present status need to be included in the records.
    All other students will be automatically marked as present.
    """

    # ✅ Only TEACHER can record attendance
    if current_user["role"] != SchoolUserRole.TEACHER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only teachers are allowed to record attendance"
        )
    
    teacher_id = uuid.UUID(current_user["user_id"])
    school_id = current_user["school_id"]

    # Get class details and verify access
    class_result = await db.execute(
        select(SchoolClass).where(
            and_(
                SchoolClass.id == attendance_data.class_id,
                SchoolClass.school_id == school_id
            )
        )
    )
    school_class = class_result.scalar_one_or_none()
    if not school_class:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Class not found or access denied"
        )
    
    # Check if date is older than 48 hours from current time
    if attendance_data.date < (datetime.utcnow().date() - timedelta(days=1.5)):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Attendance cannot be recorded for dates older than 48 hours."
        )

    # Optional: If SchoolClass has a `teacher_id`, verify ownership
    if school_class.class_teacher_id != teacher_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You are not the assigned teacher for this class"
        )

    # Get all active students in the class
    students = await _get_class_students(db, attendance_data.class_id)
    if not students:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No active students found in this class"
        )

    # Create a map of student_id to attendance input data
    attendance_input_map = {
        record.student_id: record for record in attendance_data.records
    }

    # Get all existing attendance records for the class and date in a single query
    existing_attendance_result = await db.execute(
        select(Attendance).where(
            and_(
                Attendance.school_id == school_id, 
                Attendance.class_id == attendance_data.class_id,
                Attendance.date == attendance_data.date
            )
        )
    )
    existing_attendance_map = {
        record.student_id: record for record in existing_attendance_result.scalars()
    }

    attendances = []
    for student in students:
        record = attendance_input_map.get(student.id)

        if record:
            status_val = record.status
            arrival_time = record.arrival_time
            notes = record.notes
        else:
            status_val = AttendanceStatus.PRESENT
            arrival_time = None
            notes = None

        existing = existing_attendance_map.get(student.id)
        if existing:
            existing.status = status_val
            existing.arrival_time = arrival_time
            existing.notes = notes
            existing.recorded_by = teacher_id
            attendances.append(existing)
        else:
            new_attendance = Attendance(
                school_id=school_id,
                class_id=attendance_data.class_id,
                date=attendance_data.date,
                student_id=student.id,
                status=status_val,
                recorded_by=teacher_id,
                arrival_time=arrival_time,
                notes=notes
            )
            db.add(new_attendance)
            attendances.append(new_attendance)

    await db.commit()

    # # Prepare student + teacher names
    # student_ids = [student.id for student in students]
    # users_result = await db.execute(
    #     select(SchoolUser).where(
    #         or_(
    #             SchoolUser.id.in_(student_ids),
    #             SchoolUser.id == teacher_id
    #         )
    #     )
    # )
    # users = {user.id: user for user in users_result.scalars()}

    # response_attendances = []
    # for attendance in attendances:
    #     student = users.get(attendance.student_id)
    #     recorder = users.get(attendance.recorded_by)

    #     response_attendances.append(AttendanceOut(
    #         id=attendance.id,
    #         student_id=attendance.student_id,
    #         student_name=student.name if student else "Unknown",
    #         status=attendance.status,
    #         arrival_time=attendance.arrival_time,
    #         notes=attendance.notes,
    #         recorded_by=attendance.recorded_by,
    #         recorded_by_name=recorder.name if recorder else "Unknown",
    #         created_at=attendance.created_at.isoformat()
    #     ))

    # return DailyAttendanceResponse(
    #     class_id=school_class.id,
    #     class_name=f"{school_class.class_name} {school_class.section}",
    #     date=attendance_data.date,
    #     attendances=response_attendances
    # )
    return {"message": f"Attendance recorded successfully for {attendance_data.date}"}


@router.get("/export-excel")
async def export_attendance_excel(
    class_id: str,
    from_date: str = None,
    to_date: str = None,
    db: AsyncSession = Depends(get_db),
    # current_user=Depends(get_current_user)
):
    try:
        from_dt = datetime.strptime(from_date, "%Y-%m-%d") if from_date else datetime.today() - timedelta(days=30)
        to_dt = datetime.strptime(to_date, "%Y-%m-%d") if to_date else datetime.today()

        # Fetch students
        result = await db.execute(
            select(SchoolUser)
            .where(SchoolUser.class_id == class_id)
            .order_by(SchoolUser.id)
            # .order_by(SchoolUser.roll_no.asc())
        )
        students = result.scalars().all()

        if not students:
            raise HTTPException(status_code=404, detail="No students found for this class")

        # Fetch attendance records
        attendance_result = await db.execute(
            select(Attendance)
            .where(
                Attendance.class_id == class_id,
                Attendance.date >= from_dt,
                Attendance.date <= to_dt
            )
        )
        attendance_data = attendance_result.scalars().all()

        # Map: {(student_id, date): status}
        attendance_map = {
            (record.student_id, record.date): record.status
            for record in attendance_data
        }

        all_dates = sorted(list(set(rec.date for rec in attendance_data)))

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header row
        base_headers = ["Roll No.", "Student Name", "Total Days", "Present", "Absent", "Leave", "Attendance %"]
        # Cross-platform safe formatting for Excel headers
        date_headers = [d.strftime("%d-%b").lstrip("0").replace(" 0", " ") for d in all_dates]

        headers = base_headers + date_headers

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add student data rows
        total_days = len(all_dates)
        class_present = 0
        class_absent = 0
        class_leave = 0

        for student in students:
            present = absent = leave = 0
            student_id = str(student.id) if isinstance(student.id, uuid.UUID) else student.id
            row = [student_id, student.name, total_days]

            # Count statuses
            for date in all_dates:
                status = attendance_map.get((student.id, date), "N/A")
                if status == "P":
                    present += 1
                elif status == "A":
                    absent += 1
                elif status == "L":
                    leave += 1

            total_recorded = present + absent + leave
            perc = (present / total_days) * 100 if total_days else 0
            row.extend([present, absent, leave, f"{perc:.1f}%"])

            # Append status for each date
            for date in all_dates:
                status = attendance_map.get((student.id, date), "N/A")
                row.append(status)

            ws.append(row)

            # Update class summary
            class_present += present
            class_absent += absent
            class_leave += leave

        # Summary row at bottom
        summary_row_start = len(students) + 3
        ws[f"A{summary_row_start}"] = "Class Summary"
        ws[f"A{summary_row_start}"].font = Font(bold=True)

        ws[f"A{summary_row_start + 1}"] = "Total Students"
        ws[f"B{summary_row_start + 1}"] = len(students)

        ws[f"A{summary_row_start + 2}"] = "Total Days"
        ws[f"B{summary_row_start + 2}"] = total_days

        ws[f"A{summary_row_start + 3}"] = "Total Present"
        ws[f"B{summary_row_start + 3}"] = class_present

        ws[f"A{summary_row_start + 4}"] = "Total Absent"
        ws[f"B{summary_row_start + 4}"] = class_absent

        ws[f"A{summary_row_start + 5}"] = "Total Leave"
        ws[f"B{summary_row_start + 5}"] = class_leave

        # Pie Chart
        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=summary_row_start + 3, max_row=summary_row_start + 5)
        data = Reference(ws, min_col=2, min_row=summary_row_start + 3, max_row=summary_row_start + 5)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        chart.title = "Class Attendance Distribution"
        ws.add_chart(chart, f"E{summary_row_start + 1}")

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        return FileResponse(tmp_path, filename="attendance_report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")